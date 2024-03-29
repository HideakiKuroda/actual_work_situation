from flask import flash, send_file
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta,time
import calendar
import jpholiday
import os
import shutil
import time
from copy import copy
import logging
import glob

# ロガーの設定
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

# ファイルハンドラーの設定
file_handler = logging.FileHandler('utility.log', encoding='utf-8')
file_handler.setLevel(logging.INFO)

# フォーマッターの設定
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
file_handler.setFormatter(formatter)

# ロガーにハンドラーを追加
logger.addHandler(file_handler)

def fill_dates_and_weekdays(workbook, year_month,template_path):
    sheet = workbook.active
    start_date = datetime.strptime(year_month + "-01", "%Y-%m-%d")
    _, last_day = calendar.monthrange(start_date.year, start_date.month)
    if template_path == 'work_record.xlsx':
        sheet['C3'].value = start_date.strftime('%Y年%m月')
    elif template_path == 'crew_shift_template.xlsx':   
        sheet['B2'].value = start_date.strftime('%Y年%m月') + ' 勤務表'
    
    # 曜日のインデックスを日本語の曜日にマッピング
    weekdays_jp = ["月", "火", "水", "木", "金", "土", "日"]
    
    for i in range(last_day):
        current_date = start_date + timedelta(days=i)
        date_cell = f'A{i + 7}'  # A7から開始
        weekday_cell = f'B{i + 7}'  # B7から開始
        work_cell = f'C{i + 7}'
        # 日付と日本語の曜日をセルに入力
        sheet[date_cell] = current_date.day
        sheet[weekday_cell] = weekdays_jp[current_date.weekday()]  # 曜日を日本語で表示
        sheet[work_cell] = '休日'
        sheet[work_cell].font = Font(color="FF0000")

        # 土曜日は青、日曜日と祝日は赤に設定
        if current_date.weekday() == 5:  # 土曜日
            sheet[date_cell].font = Font(color="0000FF")
            sheet[weekday_cell].font = Font(color="0000FF")
        elif current_date.weekday() == 6 or jpholiday.is_holiday(current_date):  # 日曜日または祝日
            sheet[date_cell].font = Font(color="FF0000")
            sheet[weekday_cell].font = Font(color="FF0000")
            
def generate_new_filename(base_path):
    # ファイルの基本名と拡張子を分離
    base, extension = os.path.splitext(base_path)
    counter = 1  # 連番の開始

    # 新しいファイル名を生成
    new_file_path = f"{base}({counter}){extension}"

    # 生成したファイル名が既に存在する場合は、連番を増やして再試行
    while os.path.exists(new_file_path):
        flash('ファイルが既に存在します！') 
        counter += 1
        new_file_path = f"{base}({counter}){extension}"

    return new_file_path

def show_dailyreports(monthSelect):
    # カレントディレクトリの一段前のディレクトリからのパスを取得
    base_dir = os.path.join(os.getcwd(), os.pardir, "attendance_management/dailyWorkReports")
    # 絶対パスに変換
    abs_base_dir = os.path.abspath(base_dir)
    
    # 指定されたディレクトリからファイルの一覧を取得
    try:
        files = [f for f in os.listdir(abs_base_dir) if os.path.isfile(os.path.join(abs_base_dir, f))]
    except FileNotFoundError:
        files = []
    
    # session['monthSelect']をファイル名に含むファイルのみをフィルタリング
    filtered_files = [file for file in files if monthSelect in file]
    filtered_files_path = [os.path.join(abs_base_dir, file) for file in files if monthSelect in file]
    return filtered_files

def move_files_with_retry(selected_files, new_folder_path, abs_base_dir,max_retries=5, wait_seconds=1):
    os.makedirs(new_folder_path, exist_ok=True)
    failed_files = []  # 移動に失敗したファイルを追跡するリスト
    for file_name in selected_files:
        file_path = os.path.join(abs_base_dir, file_name)
        success = False  # このファイルの移動成功フラグ
        for attempt in range(max_retries):
            try:
                shutil.move(file_path, os.path.join(new_folder_path, file_name))
                success = True  # 移動成功
                break  # 成功したらループを抜ける
            except Exception as e:
                print(f"移動に失敗しました: {file_path} -> {new_folder_path}, エラー: {e}")
                time.sleep(wait_seconds)  # 少し待ってからリトライ
        if not success:
            failed_files.append(file_name)  # 移動に失敗したファイルをリストに追加
    if failed_files:
        print(f"以下のファイルの移動に失敗しました: {failed_files}")
        return failed_files  # 一つでも失敗があればFalseを返す
    return True  # すべて成功した場合はTrueを返す

def convert_category(value):
    if value == '臨時出勤':
        return '臨出'
    elif value == '当直明け':
        return '明け'
    else:
        return value

#日報のデータを勤務ジッタ表に転記する
def copy_data_to_work_record(selected_files, work_record_file):
    target_dir = "work_records"
    work_record_path = os.path.join(target_dir, work_record_file)
    wb_work_record = load_workbook(work_record_path)
    ws_work_record = wb_work_record.active
    file_move_destinations = {}
    default_font = Font(color="000000") 
    base_dir = os.path.join(os.getcwd(), os.pardir, "attendance_management/dailyWorkReports")
    # base_dir2 = os.path.join(os.getcwd(), os.pardir, "attendance_management/end_dailyWorkReports")
    abs_base_dir = os.path.abspath(base_dir)
    # abs_base_dir2 = os.path.abspath(base_dir2)
    failed_moves = []
    for file_name in selected_files:
        file_path = os.path.join(abs_base_dir, file_name)
        wb_report = load_workbook(file_path, data_only=True)
        ws_report = wb_report.active
        # 勤務実態表と作業日報の日付が合っているかチェック
        date_str = ws_work_record['C3'].value # 勤務実態表の年月
        # 年と月を抽出するために文字列を解析する
        year, month = map(int, date_str[:-1].split('年'))
        # 日を手動で設定（ここでは月の最初の日として1日を使用）
        date_work_record = datetime(year, month, 1)
        date_report = datetime.strptime(ws_report['B4'].value, "%Y年%m月%d日")  # 作業日報の日付
        if date_work_record.strftime("%Y-%m") == date_report.strftime("%Y-%m"):
            # 日付が一致する行を見つける
            for row in range(7, 38):  # A7～A37
                if ws_work_record[f'A{row}'].value == date_report.day:
                    # データをコピー
                    ws_work_record[f'C{row}'].value = convert_category(ws_report['F4'].value)
                    ws_work_record[f'C{row}'].font = default_font
                    ws_work_record[f'D{row}'].value = ws_report['C22'].value
                    ws_work_record[f'F{row}'].value = ws_report['C23'].value
                    ws_work_record[f'H{row}'].value = ws_report['V13'].value
                    ws_work_record[f'J{row}'].value =  '' if ws_report['V15'].value.strftime('%H:%M')=='00:00' else ws_report['V15'].value
                    ws_work_record[f'L{row}'].value =  '' if ws_report['V16'].value.strftime('%H:%M')=='00:00' else ws_report['V16'].value
                    break
        new_folder_path = os.path.join(base_dir, date_report.strftime("%Y-%m"))        
        wb_report.close()
        wb_report.close()
       # 勤務実態表ファイルを保存
    wb_work_record.save(work_record_path)
    # time.sleep(3)  # 2秒待機
    # # 作業日報ファイルを新しいフォルダに移動
    # failed_moves = move_files_with_retry(selected_files, new_folder_path,abs_base_dir)
    # if failed_moves:
    #     flash("以下のファイルの移動に失敗しました:")
    #     for file_name in failed_moves:
    #         flash(f"{file_name} -> {new_folder_path}")
    # else:
    flash("勤務実態表ファイルの書き込みが完了しました。ファイルがダウンロードされます")       


def copy_cell(source_cell, target_cell):
    target_cell.value = source_cell.value
    target_cell.font = copy(source_cell.font)
    target_cell.border = copy(source_cell.border)
    target_cell.fill = copy(source_cell.fill)
    target_cell.number_format = source_cell.number_format
    target_cell.alignment = copy(source_cell.alignment)

def copy_dates_to_new_sheet(workbook, template_sheet_name, target_sheet_name):
    template_sheet = workbook[template_sheet_name]
    target_sheet = workbook[target_sheet_name]
    column_widths = [5.4, 5.4, 6.2, 6.2, 4.7, 4.7, 4.7, 4.7, 3.7, 3.7, 4.7, 4.7, 4.7, 4.7, 4.5, 4.5, 8.1]
    for i, width in enumerate(column_widths, start=1):
        col_letter = get_column_letter(i)
        target_sheet.column_dimensions[col_letter].width = width+0.7
 # セルの内容をコピー
    for row in template_sheet.iter_rows(min_row=1, max_row=39, min_col=1, max_col=17):
        for cell in row:
            target_cell = target_sheet.cell(row=cell.row, column=cell.column)
            copy_cell(cell, target_cell)
    # 結合されたセルの範囲をコピー
    for merge_cell in template_sheet.merged_cells.ranges:
        target_sheet.merge_cells(str(merge_cell))

        # 行の高さをコピー（特定の範囲）
    min_row, max_row = 1, 39
    for row_idx in range(min_row, max_row + 1):
            if row_idx in template_sheet.row_dimensions:
                target_sheet.row_dimensions[row_idx].height = template_sheet.row_dimensions[row_idx].height

#crew_shift　の中の日付「あさか丸」乗組員勤務表の最新 を取得
def find_latest_file(month):
    files = glob.glob(f'crew_shift/{month}「あさか丸」乗組員勤務表*.xlsx')
    latest_file = max(files, key=os.path.getctime)
    return latest_file   

#勤務実態表から乗組員勤務表へのコピーを部分的に修正します
def modify_work_records(n_work_wb, work_records_wsheet):
    # n_work_wb の全シートを処理
    for sheet_name in n_work_wb.sheetnames:
        sheet = n_work_wb[sheet_name]
        for i in range(7, 38):  # C7からC37までの範囲で処理
            # I列が空白で、C列に'日勤'が記載されている行を'休'に変更し、D列を空白に
            if sheet[f'I{i}'].value is None and sheet[f'C{i}'].value == '日勤':
                sheet[f'C{i}'] = '休日'
                sheet[f'C{i}'].font = Font(color="FF0000")  # 赤色を指定
                sheet[f'D{i}'] = None

            # ここから先の既存の処理
            if sheet[f'C{i}'].value == '休日' and sheet[f'D{i}'].value is not None:
                # C列が'休日'でD列が空白でない場合、Dを空白にする
                sheet[f'D{i}'] = None
            elif sheet[f'C{i}'].value == '明け' and sheet[f'D{i}'].value is None:
                # C列が'明け'でD列が空白の場合、前の行のD列をコピー
                sheet[f'D{i}'] = sheet[f'D{i-1}'].value
                sheet[f'D{i}'].font = Font(color="000000")  # 黒色を指定

            # work_records_wsheetのC7～37を確認
            work_duty = work_records_wsheet[f'C{i}'].value if i <= 37 else None  # work_records_wsheet の範囲を超えないように調整
            if work_duty in ['休日▲', '▲休日'] and sheet[f'C{i}'].value == '休日':
                # 対応するsheetのC列が'休'であれば、'▲休'に変更
                sheet[f'C{i}'] = '▲休日'
                sheet[f'C{i}'].font = Font(color="FF0000") 

#勤務実態表から乗組員勤務表へのコピー
def copy_work_records(n_work_wb, month_str):
    # 勤務実態表ファイルを探す
    work_records_path = 'work_records'
    files = os.listdir(work_records_path)
    work_records_file = next((file for file in files if file.startswith(month_str)), None)
    if work_records_file is None:
        print(f"{month_str}で始まるファイルが見つかりません。")
        return

    work_records_wbook = load_workbook(os.path.join(work_records_path, work_records_file))
    work_records_wsheet = work_records_wbook.active

    # n_work_wb の全シートを処理
    for sheet_name in n_work_wb.sheetnames:
        sheet = n_work_wb[sheet_name]
        for i in range(7, 38):  # C7からC37まで
            duty = work_records_wsheet[f'C{i}'].value
            if duty in ['日勤', '当直']:
                # A列で一致する行をn_work_wbのシートで探す
                day = work_records_wsheet[f'A{i}'].value
                for j in range(7, 38):  # n_work_wbのシートのA7からA37まで探す
                    if sheet[f'A{j}'].value == day and sheet[f'C{j}'].value == '日勤':
                        # データをコピー
                        for col in [('D', 'E'), ('F', 'G'), ('H', 'I'), ('J', 'K'), ('L', 'M')]:
                            sheet[f'{col[1]}{j}'] = work_records_wsheet[f'{col[0]}{i}'].value
                        if duty == '当直':
                            sheet[f'C{j}'].value = '当直'
                            sheet[f'C{j}'].font = Font(color="000000")  # 黒色を指定
                            # 次の行にデータをコピーし、「明け」に設定
                            for col in [('D', 'E'), ('F', 'G'), ('H', 'I'), ('J', 'K'), ('L', 'M')]:
                                sheet[f'{col[1]}{j+1}'] = work_records_wsheet[f'{col[0]}{i+1}'].value
                            sheet[f'C{j+1}'].value = '明け'
                            sheet[f'C{j+1}'].font = Font(color="000000")  # 黒色を指定
                            break
    modify_work_records(n_work_wb, work_records_wsheet)                    
    work_records_wbook.close

def create_crew_commuting(month_str, n_work_wb):
    # 乗組員通勤費明細書のテンプレートからmonthを頭に付けたファイルを作成します
    new_filename = f"{month_str}あさか丸乗組員通勤費明細書.xlsx"
    template_path = 'commuting_allowance.xlsx'
    new_file_path = generate_new_filename(os.path.join('crew_shift', new_filename))
    workbook = load_workbook(template_path)
    template_sheet = workbook.active  # テンプレートのアクティブシートを取得

    # 乗組員勤務表でシート毎に処理します
    for sheet_name in n_work_wb.sheetnames:
        if sheet_name in workbook.sheetnames:
            n_sheet = n_work_wb[sheet_name]
            t_sheet = workbook[sheet_name]

            # 日付に基づいて往路と復路の交通費を計算する
            for i in range(7, 38):  # 日付の範囲
                duty = n_sheet[f'C{i}'].value
                if duty == '日勤':
                    # 日付けに応じたセルに○を入れる
                    day = i - 6
                    if day <= 10:
                        t_sheet[f'D{25 + day}'].value = t_sheet[f'E{25 + day}'].value = '○'
                    elif day <= 20:
                        t_sheet[f'G{15 + day}'].value = t_sheet[f'H{15 + day}'].value = '○'
                    elif day <= 31:
                        t_sheet[f'J{5 + day}'].value = t_sheet[f'K{5 + day}'].value = '○'
                elif duty == '当直':
                    day = i - 6
                    if day <= 10:
                        t_sheet[f'D{25 + day}'].value = '○'
                    elif day <= 20:
                        t_sheet[f'G{15 + day}'].value = '○'
                    elif day <= 31:
                        t_sheet[f'J{5 + day}'].value = '○'
                elif duty == '明け':
                    day = i - 6
                    if day <= 10:
                        t_sheet[f'E{25 + day}'].value = '○'
                    elif day <= 20:
                        t_sheet[f'H{15 + day}'].value = '○'
                    elif day <= 31:
                        t_sheet[f'K{5 + day}'].value = '○'

    # ファイルを保存します
    workbook.save(new_file_path)
    return new_file_path

def download_file(filepath):
    # filepath = 'work_records/sample.xlsx'  # ダウンロードさせたいファイルのパス
    return send_file(filepath, as_attachment=True)

logging.basicConfig(filename='utility.log', level=logging.INFO, 
                    format='%(asctime)s %(levelname)s:%(message)s')