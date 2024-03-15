from flask import Flask, request, render_template, flash, session,redirect, url_for, jsonify,send_file
from openpyxl import load_workbook
from openpyxl.styles import Font
from utility import fill_dates_and_weekdays,generate_new_filename,show_dailyreports,copy_data_to_work_record,\
copy_dates_to_new_sheet,find_latest_file,copy_work_records, create_crew_commuting
from datetime import datetime
import os
import logging
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.secret_key = 'secret_key8902083508##'
app.config.update(
    SESSION_COOKIE_SECURE=True,
    SESSION_COOKIE_SAMESITE='None',
)
app.config['UPLOAD_FOLDER'] = 'uploads'
# ロガーの設定
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

# ファイルハンドラーの設定
file_handler = logging.FileHandler('app.log', encoding='utf-8')
file_handler.setLevel(logging.INFO)

# フォーマッターの設定
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
file_handler.setFormatter(formatter)

# ロガーにハンドラーを追加
logger.addHandler(file_handler)

#最初の画面を開く
@app.route('/', methods=['GET', 'POST'])
def index():
    session.clear
    return render_template("index.html")

#'その月の新しい勤務実態表を作成して日付を入力
@app.route('/create_report', methods=['GET', 'POST'])
def create_report():
	session['monthSelect'] = request.form['monthSelect']
	new_filename = f"{session['monthSelect']}「あさか丸」勤務実態表.xlsx"
	template_path = 'work_record.xlsx'
	new_file_path = generate_new_filename(os.path.join('work_records', new_filename))
	workbook = load_workbook(template_path)
	#テンプレートから新しいファイルを作成、日付と曜日を入れる
	fill_dates_and_weekdays(workbook, session['monthSelect'],template_path)
    # sessionのキーに対応するリストを初期化（これはアプリケーションのどこかで行う必要があります）
	session['dailyreports'] = show_dailyreports(session['monthSelect'])
	#ファイルの保存
	workbook.save(new_file_path) 
	flash('新しい勤務実態表を作成しました! %s' % new_filename)
    # 'work_records'フォルダ内のファイルリストを取得
	files = os.listdir('work_records')
	#files_path = ['work_records/' + file for file in files]  # パスを含むファイル名のリスト
	start_date = datetime.strptime(session['monthSelect']+ "-01", "%Y-%m-%d")
	session['start_date'] = start_date.strftime('%Y年%m月') 
	session['files'] = files  # セッションにファイルリストを保存
	session['new_filename'] = os.path.basename(new_file_path)
	return redirect(url_for('show_files'))

#その月の勤務実態表を作成に入れるデータを選択する
@app.route('/show_files')
def show_files():
    files = session.get('files', [])  # セッションからファイルリストを取得
    monthSelect = session.get('start_date','') 
    selected_file = session.get('new_filename','')
    dailyreports = session.get('dailyreports','')
	
    return render_template("show_files.html", files=files, monthSelect=monthSelect, selected_file=selected_file, dailyreports=dailyreports)

@app.route('/edit_report', methods=['GET', 'POST'])
def edit_files():
    selected_files = request.form.getlist('selected_files')
    work_record_file = session.get('new_filename','')
    copy_data_to_work_record(selected_files,  work_record_file)
    target_dir = "work_records"
    work_record_path = os.path.join(target_dir, work_record_file)
    return render_template("index.html", filename=work_record_path)

#乗組員の勤務表を作成（テンプレートからコピー作成、その月の日付を入力し、乗組員ごとのフォルダを作成）
@app.route('/create_crew_shift', methods=['GET', 'POST'])
def create_crew_shift():
    if request.method == 'POST':
        try:
            data = request.get_json()
            month  = data.get('month', None)
            session['monthSelect'] = month
            logger.info('monthSelect: %s', month) 
            new_filename = f"{session['monthSelect']}「あさか丸」乗組員勤務表.xlsx"
            template_path = 'crew_shift_template.xlsx'
            new_file_path = generate_new_filename(os.path.join('crew_shift', new_filename))
            workbook = load_workbook(template_path)
            # テンプレートから新しいファイルを作成し、日付と曜日を入れる
            fill_dates_and_weekdays(workbook, session['monthSelect'], template_path)
            # 'work_records'ディレクトリ内のファイルを確認
            work_records_path = 'work_records'
            month_str = session['monthSelect']
            files = os.listdir(work_records_path)
            # logger.info('month_str: %s', month_str) 
            # 対応する「勤務実態表」があるかどうかを確認
            if not any(file.startswith(month_str) for file in files):
                # logger.info('files_None: %s', files) 
                return jsonify({'error': "対応する「勤務実態表」がありません"}), 404
            # 'name.txt'にリストされた名前でシートを作成
            with open('name.txt', 'r', encoding='utf-8') as names_file:
                names = names_file.read().splitlines()
                for name in names:
                    if name not in workbook.sheetnames:
                        sheet = workbook.create_sheet(title=name)
                        # 新しく作成したシートに日付や曜日などをコピーする
                        copy_dates_to_new_sheet(workbook, 'name', name)
                        sheet['D3'] = name
            # logger.info('names: %s', names)             
            # # 元にあった'name'という名前のシートを削除
            if 'name' in workbook.sheetnames:
                std = workbook['name']
                workbook.remove(std)
            
            workbook.save(new_file_path)
            return jsonify({'message': "乗組員の勤務表が正常に作成されました。",'redirect': '/sheet_selector'}), 200     
        except Exception as e:
            app.logger.error(f'Error occurred: {e}')
            return jsonify({'error': "エラーが発生しました！処理をやり直してください。"}), 404

@app.route('/sheet_selector', methods=['GET', 'POST'])
def sheet_selector():
    month_str =  session['monthSelect']
    return render_template("sheet_selector.html",month_str=month_str)

#あさか丸勤務表のファイルとシートを選択
@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'})
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'})
    if file:
        filename = secure_filename(file.filename)
        session['secure_file'] = filename
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        session['upload_filename'] = filename
        return jsonify({'success': 'File uploaded successfully', 'filename': filename})
    
#選択したファイルのシートをリスト化
@app.route('/sheets', methods=['POST'])
def get_sheets():
    filename = request.json['filename']
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    workbook = load_workbook(filepath)
    sheets = workbook.sheetnames
    return jsonify(sheets)

    # logger.info('sheetName: %s',  a_work_sheet_name) 
#crew_shift　の中の日付「あさか丸」乗組員勤務表 を作成
@app.route('/copy_data_to_n_work_file', methods=['POST', 'GET'])
def copy_data_to_n_work_file():
    a_work_sheet_name = request.form.get('sheetName')
    # A勤務表を開く
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], session['secure_file'])
    a_work_wb = load_workbook(filepath)  
    a_work_sheet = a_work_wb[a_work_sheet_name]

    # N勤務ファイルを開く
    month =  session['monthSelect']
    filename = find_latest_file(month)
    n_work_wb = load_workbook(filename)
    
    # 対象のセル範囲を指定
    cell_ranges = ['E3', 'H3', 'K3', 'N3', 'Q3', 'T3', 'W3', 'Z3']
    target_ranges = ['E5:E35', 'H5:H35', 'K5:K35', 'N5:N35', 'Q5:Q35', 'T5:T35', 'W5:W35', 'Z5:Z35']
    
    # Textが空のセルが出たらそこで終わるように調整
    for i, cell in enumerate(cell_ranges):
        if not a_work_sheet[cell].value:
            cell_ranges = cell_ranges[:i]
            target_ranges = target_ranges[:i]
            break

    for cell, target_range in zip(cell_ranges, target_ranges):
        cell_value = a_work_sheet[cell].value
        if cell_value is not None:
            sheet_name = cell_value[:2]
            logger.info('sheet_name: %s', sheet_name)
        else: sheet_name ='' # 頭2文字を取得
        matching_sheets = [s for s in n_work_wb.sheetnames if s.startswith(sheet_name)]

        if matching_sheets:
            n_sheet = n_work_wb[matching_sheets[0]]
            logger.info('n_sheet: %s', n_sheet)
            values = [cell.value for row in a_work_sheet[target_range] for cell in row]
            for i, val in enumerate(values, start=7):  # C7から始める
                if val in ('船長', '一航士', '機関長', '一機士', '甲板員', '機関員'):
                    cell = f'd{i}'
                    n_sheet[cell].value = val
                    n_sheet[cell].font = Font(color="000000")  # 黒色を指定
                    cell = f'C{i}'
                    n_sheet[cell].value = '日勤'
                    n_sheet[cell].font = Font(color="000000")  # 黒色を指定
                    # フォントを黒色に設定
    a_work_wb.close
    n_work_wb.save(filename)  # 更新されたファイルの保存先

    copy_work_records(n_work_wb, month)
    new_file_path = create_crew_commuting(month, n_work_wb)
    n_work_wb.save(filename)  # 更新されたファイルの保存先
    flash('乗組員の勤務表が正常に作成されました。ファイルがダウンロードされます。') 
    return render_template("index.html", filename=filename,new_file_path=new_file_path)
    # return redirect(url_for('download_file', filename=filename))

@app.route('/download/<filename>')
def download_file(filename):
    # セキュリティの観点から、実際のファイルパスを慎重に扱い、
    # 直接ユーザー入力をファイルパスに使用しないようにしてください。
    filepath = filename
    return send_file(filepath, as_attachment=True)


logging.basicConfig(filename='app.log', level=logging.INFO, 
                    format='%(asctime)s %(levelname)s:%(message)s')

if __name__ == '__main__':
    app.run(debug=True)